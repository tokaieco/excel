from openpyxl.styles import Alignment
import streamlit as st
import pandas as pd
# datetimeモジュールを使用
import datetime
from dateutil.relativedelta import relativedelta  # to add days or years

import time

import openpyxl
import pprint
from openpyxl import Workbook
import base64
from PIL import Image

t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
# ログイン
col1, col2 = st.columns((4, 2))
with col1:
    # st.text_input("氏名")
    option = st.selectbox(
        '氏名',
        ('吉田', '村上', '岡本', '牧', '源田', '中村', 'ヌートバー', '近藤', '大谷'))

st.write('ログイン:', option)
with col2:
    st.text_input("パスワード")

# 西暦を2桁に
# d = now.strftime('%y/%m/%d %H:%M:%S')
d = now.strftime('%y/%m/%d %H:%M')
st.title('予定表')
# st.text(d)
df = pd.read_excel('予定表1.xlsx', sheet_name='Sheet1')
# df= df1.rename(index=lambda s: s.replace(" ",""))
# df.sort_values('品番', ascending=False)
# df['品番'] = df['品番'].astype(str)

# view_columns = ['品番', '工程順位']
# ラベル指定
# df.loc[:, ['品番'.isin('197'), '品名', '工程順位', '受注数量']]

# df1 = df[df['品番'] == 197]
# hyo = df1.loc[:, ['品名', '工程順位', '受注数量', '開始数量', 'NG数', '終了数量', '備考']]
# サイド
add_selectbox = st.sidebar.selectbox(
    "NG数", (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10))


# cols1, _ = st.beta_columns((1, 2))  # To make it narrower
# format = 'MMM DD, YYYY'  # format output
# start_date, end_date = st.date_input('start date  - end date :', [])
# if start_date < end_date:
#    pass
# else:
#    st.error('Error: End date must fall after start date.')

# mask = (df['CREATEDDATE'] > start_date) & (df['CREATEDDATE'] <= end_date)


# ボタン１
# st.text_input("Message", key="text_input")  # (A)


# def change_value():
#    st.session_state["text_input"] = "Hello, World"  # (B)


# st.button("Click", on_click=change_value)  # (C)

# ボタン２
# if st.button(label='click me!'):
#    st.write('Thank you')

# ボタン3
dt_now5 = time.strftime('%y.%m.%d %H:%M')
# dt_now3 = time.strftime('%Y/%m/%d %H.%M')
dt_now3 = time.strftime('%y.%m.%d %H:%M')
dt_now4 = time.strftime('%y.%m.%d %H:%M')
# if st.button(label='開始!'):
#    st.write(dt_now3)
st.session_state.key = dt_now3
# カラムごとに幅を比率で指定 (この場合は１：２：４：２)
# col1, col2, col3, col4, col5 = st.columns((4, 1, 1, 1, 1))


##ここから第4回の内容##
##国名の列にある全ての国をリスト化する##
erea_list = df['品番'].unique()

##国名のセレクトボックスを作成する##
selected_erea = st.sidebar.selectbox(
    '表示する品番を選択：',
    erea_list
)

##セレクトされた国でデータフレームの中身をフィルタリングする##
df2 = df[df['品番'] == selected_erea]
hyo2 = df2.loc[:, ['品番', '品名', '得意先', '工程順位',
                   '受注数量', '開始数量', 'NG数', '氏名', '備考']]
st.dataframe(hyo2)

# excel記入


# book = openpyxl.Workbook()

# active_sheet = book.active
# active_sheet['AA2'] = 1
# active_sheet['AA3'] = 10
# active_sheet['AA4'] = 100
# active_sheet['AA5'] = '=sum(AA2,AA3,AA4)'

# book.save('sample.xlsx')


#col3, col4 = st.columns((8, 2))
# with col3:
#    hyo2
# with col4:
# st.text_input("その他")
#btn = st.button("作業手順書", key=0)

# 動画

if 'count' not in st.session_state:
    st.session_state["count"] = 0

# if st.button("作業手順書", key=0):
# if btn:
col1, col2 = st.columns((7, 2))
with col1:
    if st.button("作業手順書", key=9):
        video_file = open('doga.mp4', 'rb')
        video_bytes = video_file.read()
        st.video(video_bytes)

with col2:
    st.button('Close 手順書', key='10')









# pdf


#def show_pdf(file_path):
#    with open(file_path, "rb") as f:
#        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
#    pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="800" height="800" type="application/pdf"></iframe>'
#    st.markdown(pdf_display, unsafe_allow_html=True)

#jpg
image = Image.open('test.jpg')

#st.image(image, caption='図面', use_column_width=True)

# show_pdf('ボンゴ牛刀.pdf')

col1, col2 = st.columns((7, 2))
with col1:
    if st.button('図面', key='7'):
        st.image(image, caption='図面', use_column_width=True)
        #show_pdf('ボンゴ牛刀.pdf')
with col2:
    st.button('Close 図面', key='8')



# サイドバー
with st.sidebar:
    # ボタン3
    dt_now5 = time.strftime('%y.%m.%d %H:%M')

    st.text_input("開始時間", key="text_input2")  # (A)

    def change_value2():
        st.session_state["text_input2"] = dt_now5  # (B)

    st.button("Click", on_click=change_value2, key=5)  # (C)

    # ボタン4
    dt_now6 = time.strftime('%y.%m.%d %H:%M')

    st.text_input("終了時間", key="text_input3")  # (A)

    def change_value3():
        st.session_state["text_input3"] = dt_now6  # (B)

    st.button("Click", on_click=change_value3, key=6)  # (C)

    # 備考追加
    text = st.text_input("気になることを入力してください")

    if 'text_list' not in st.session_state:
        st.session_state["text_list"] = []

    col1, col2 = st.columns(2)

    with col1:
        if st.button("追加", key=2):
            st.session_state["text_list"].append(text)

    with col2:
        if st.button("削除", key=3):
            st.session_state["text_list"].remove(text)

    for output_text in st.session_state["text_list"]:
        st.write("", output_text)

    # 書込み
    # ボタン5
    # dt_now7 = time.strftime('%y.%m.%d %H:%M')

    # st.text_input("終了", key="text_input4")  # (A)

    def change_value4():
        # st.session_state["text_input4"] = str('123')  # (B)
        # st.button("Click", on_click=change_value4, key=7)  # (C)
        # df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
        #                  index=['2', '3', '4'], columns=['開始時間', '終了時間', '差'])
        # with pd.ExcelWriter('data/dst/pandas_to_excel.xlsx', mode='a') as writer:
        #    df.to_excel(writer, sheet_name='new_sheet1')

        #active_sheet = hyo2
        #active_sheet['AA10'] = dt_now6
        #df['10', '開始時間'] = 'Hello'

        # with pd.ExcelWriter('予定表1.xlsx', engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
        #    # with pd.ExcelWriter('予定表1.xlsx', engine="openpyxl", mode='a') as writer:
        #   df.to_excel(writer, sheet_name='Sheet2')

        #wb = openpyxl.Workbook()
        import openpyxl as px  # openpyxlモジュールをpxとしてインポートする
        # sample.xlsx を読み込み、Workbookオブジェクトを取得してwbとする
        wb = px.load_workbook('予定表1.xlsx')
        ws = wb.worksheets[0]  # sheeet=book['sheet1']
        ws["AB2"] = dt_now5
       # ws["AC2"].value = "World"
        ws["AC2"] = dt_now6
        #ws["11", '開始時間'] = "Hello3"
        ws["Z2"] = text
        #ws["AE2"] = ws["H2"]
        ws["AE2"] = ws["H2"].value
        ws["AG2"] = add_selectbox
        ws["AF2"] = ws["AE2"].value-ws["AG2"].value

        wb.save("予定表1.xlsx")
        wb.close()		        # Workbookオブジェクトを閉じる
      # ワークブックの読み込み
       # from openpyxl import load_workbook
       #wb = load_workbook(df)
        # ws = wb['Sheet1']  # ワークシートを指定
        # ws=df.active
        #ws = wb.active
        # ws['A1'] = 'Hello from Python'
        #df['AA3'] = 'Hello'

        # Xdf['Sheet1'].save('myworkbook.xlsx')  # overwrite myworkbook.xlsx

    st.button("終了", on_click=change_value4)
    # df.to_excel('予定表1.xlsx', sheet_name='Sheet1')
    # ws['B14'].value = 10.0
    # print(ws['B14'].value)
    # df.to_excel('仕様書1.xlsx', startrow=2, startcol=27)


# セレクト
# skill_option = st.selectbox(
#    'Which skill do you most want to learn?',
#    ('Java', 'Python', 'C', 'PHP', 'C++', 'Javascript', 'HTML', 'Other'))
# st.write('You selected:', skill_option)
# st.time_input
# st.title('セレクトボックス')
st.header('項目選択')

stock = st.selectbox(label="項目を選んでください",
                     options=df.columns)


st.line_chart(df[[stock]])
# 書込み
# df2 = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
#                   index=['2', '3', '4'], columns=['開始時間', '終了時間', '差'])
# df.to_excel('予定表1.xlsx', sheet_name='Sheet1')
# ws['B14'].value = 10.0
# print(ws['B14'].value)
# df.to_excel('仕様書1.xlsx', startrow=2, startcol=27)
# with pd.ExcelWriter('予定表1.xlsx', engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
#    df.to_excel(writer, sheet_name='Sheet1')
